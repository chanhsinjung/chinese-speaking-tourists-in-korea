import os
import time
from urllib.parse import urlparse

import openpyxl
from selenium.common import NoSuchElementException
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait


def comm(driver: WebDriver):
    def need_get(ele: WebElement):
        try:
            r = ele.find_element(By.CLASS_NAME, "r")
            if not r.get_attribute('style').find("display: none;"):
                return True
            else:
                return False
        except NoSuchElementException:
            return False

    def get_address(ele: WebElement):
        try:
            return ele.find_element(By.CLASS_NAME, "address").text
        except NoSuchElementException:
            return ""

    def get_by_class(ele: WebElement, class_name: str):
        try:
            return ele.find_element(By.CLASS_NAME, class_name)
        except NoSuchElementException:
            return None

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    driver.implicitly_wait(1)

    comment_list = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "gl-poi-detail_comment-list"))
    )
    time.sleep(2)

    y = 4500
    driver.execute_script("window.scrollTo(arguments[0], arguments[1]);", 0, y)
    driver.implicitly_wait(1)

    poi_page_title = (driver.find_element(By.CLASS_NAME, "poi-page-title")
                      .find_element(By.CLASS_NAME, "basicName").text)
    active = (driver.find_element(By.CLASS_NAME, "gl-poi-detail_page")
              .find_element(By.CLASS_NAME, "gl-cpt-pager")
              .find_element(By.CLASS_NAME, "active").text)
    next_btn = (driver.find_element(By.CLASS_NAME, "gl-poi-detail_page")
                .find_element(By.CLASS_NAME, "gl-cpt-pagination")
                .find_element(By.CLASS_NAME, "btn-next"))
    if poi_page_title + "-" + active in get_list:
        return poi_page_title, active, next_btn

    review_items = comment_list.find_elements(By.CLASS_NAME, "review-item")
    for item in review_items:
        y += 1000
        driver.execute_script("window.scrollTo(arguments[0], arguments[1]);", 0, y)
        driver.implicitly_wait(1)
        if need_get(item):
            ReviewUser_dom = WebDriverWait(item, 5).until(
                EC.presence_of_element_located((By.CLASS_NAME, "review-user-name"))
            )
            ReviewUser = ReviewUser_dom.text

            Score = get_by_class(item, "score-box")
            if Score is None:
                Score = ""
            else:
                Score = get_by_class(Score, "Score").text

            ScoreName = get_by_class(item, "score-name")
            if ScoreName is None:
                ScoreName = ""
            else:
                ScoreName = ScoreName.text

            Review = (item.find_element(By.CLASS_NAME, "gl-poi-detail_comment-content")
                      .find_elements(By.XPATH, "./div"))

            Like = item.find_element(By.CLASS_NAME, "ThumbsUpStyle-sc-1vbhow2-0").text
            create_time = item.find_element(By.CLASS_NAME, "create-time").text

            IdName = ""
            UserName = ""
            Nationality = ""
            TotalPost = 0
            Fan = 0
            Following = 0
            if ReviewUser != "匿名用戶":
                user = item.find_element(By.CLASS_NAME, "review-user-view")
                driver.execute_script("arguments[0].click();", user)
                driver.implicitly_wait(10)
                if len(driver.window_handles) >= 2:
                    driver.switch_to.window(driver.window_handles[1])
                    current_url = driver.current_url
                    parsed_url = urlparse(current_url).path.split("/")
                    IdName = parsed_url[len(parsed_url) - 1]

                    infor = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "infor"))
                    )
                    UserName = WebDriverWait(infor, 10).until(
                        EC.presence_of_element_located((By.CLASS_NAME, "name"))
                    ).text
                    Nationality = get_address(infor)
                    info_nums = infor.find_elements(By.CLASS_NAME, "info-num")
                    TotalPost = info_nums[0].text
                    Fan = info_nums[1].text
                    Following = info_nums[2].text

                    driver.execute_script("window.close()")
                    driver.switch_to.window(driver.window_handles[0])

            print(poi_page_title, active, ReviewUser)
            sheet.append([poi_page_title,
                          active,
                          ReviewUser,
                          Score,
                          ScoreName,
                          Review[1].text,
                          Like,
                          create_time,
                          IdName,
                          UserName,
                          Nationality,
                          TotalPost,
                          Fan,
                          Following])

    return poi_page_title, active, next_btn


def get_comm(driver: WebDriver):
    poi_page_title, active, btn_next = comm(driver)
    excel.save(file)
    get_list[poi_page_title + "-" + active] = active

    # if int(active) >= 2:
    #     return

    if btn_next.get_attribute('class').find('disabled') == -1:
        driver.execute_script("arguments[0].click();", btn_next)
        driver.implicitly_wait(10)
        get_comm(driver)


def spider(src: str):
    with Chrome() as driver:
        driver.get(src)
        get_comm(driver)
        time.sleep(2)
        driver.close()


if __name__ == '__main__':
    url_list = [
        # 'https://hk.trip.com/travel-guide/attraction/seoul/gwanghwamun-square-90204/?rankingId=100000001628&locale=zh-HK&curr=TWD'
        # 'https://hk.trip.com/travel-guide/attraction/seoul/cheongdam-dong-10800198/?rankingId=100000001628&locale=zh-HK&curr=TWD'
        # 'https://hk.trip.com/travel-guide/attraction/seoul/leeum-museum-of-art-90379/?rankingId=100000001628&locale=zh-HK&curr=TWD'
        # 'https://hk.trip.com/travel-guide/attraction/seoul/bongeunsa-88521/?rankingId=100000001628&locale=zh-HK&curr=TWD'
        # 'https://hk.trip.com/travel-guide/attraction/seoul/gocheok-sky-dome-30464869/?rankingId=100000001628&locale=zh-HK&curr=TWD'
        # 'https://hk.trip.com/travel-guide/attraction/seoul/gyeongbokgung-palace-78910/?rankingId=100000001628&locale=zh-HK&curr=TWD'
        # 'https://hk.trip.com/travel-guide/attraction/seoul/coex-aquarium-10521755/?rankingId=100000001628&locale=zh-HK&curr=TWD'
    ]

    file = "./data.xlsx"
    get_list = {}
    if os.path.exists(file):
        excel = openpyxl.load_workbook(file)
        sheet = excel.active
        last_row_index = sheet.max_row
        if last_row_index > 1:
            for row in range(2, last_row_index + 1):
                Place = sheet.cell(row, 1).value
                Page = sheet.cell(row, 2).value
                get_list[Place + "-" + Page] = Page

    else:
        excel = openpyxl.Workbook()
        sheet = excel.active
        sheet.title = "spider"
        sheet.append(
            ["Place", "Page",
             "ReviewUser",
             "Score",
             "ScoreName",
             "Review",
             "Like",
             "create_time",
             "IdName",
             "Username",
             "Nationality",
             "TotalPost",
             "Fan",
             "Following"])

    for url in url_list:
        spider(url)

    excel.save(file)
    excel.close()
